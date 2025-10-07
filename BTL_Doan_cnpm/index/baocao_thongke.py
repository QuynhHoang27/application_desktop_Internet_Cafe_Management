
from PyQt6 import QtCore, QtGui, QtWidgets, uic
from PyQt6.QtWidgets import *
from PyQt6.QtWidgets import QMenu, QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
import sys
import mysql.connector
from PyQt6.QtCore import QDate


class baocao_thongke(QMainWindow):
    def __init__(self,ten_tai_khoan):
        super(baocao_thongke, self).__init__()
        uic.loadUi('/Users/hoangquynh/BTL_DOAN_CNPM/interface/interface_baocao_thongke.ui', self)
        self.date_ngay.setCalendarPopup(True)   
        self.date_ngay.setDisplayFormat("yyyy-MM-dd")
        self.date_thang.setCalendarPopup(True) 
        self.date_thang.setDisplayFormat("yyyy-MM")  
        self.date_ngay.setDate(QDate.currentDate())
        self.date_thang.setDate(QDate.currentDate())
        style_date = """
            QDateEdit {
                padding: 6px 10px;
                border: 1px solid #aaa;
                border-radius: 8px;
                min-width: 120px;
                background: white;
                color: black;
                font-size: 14px;
            }
            QDateEdit QAbstractItemView {
                background: white;
                color: black;
                selection-background-color: #5e60ce;
                selection-color: white;
            }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 25px;
                border-left: 1px solid #aaa;
            }
            QDateEdit::down-arrow {
                width: 12px;
                height: 12px;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #5e60ce;   /* nền thanh tiêu đề */
            }
            QCalendarWidget QToolButton {
                color: white;                /* màu chữ tháng/năm + nút mũi tên */
                font-size: 14px;
                icon-size: 16px;
                background: transparent;
            }
            QCalendarWidget QToolButton:hover {
                background-color: #4ea8de;
                border-radius: 4px;
            }
        """
        self.date_ngay.setStyleSheet(style_date)
        self.date_thang.setStyleSheet(style_date)



        # Làm đẹp nút lọc
        self.btn_loc.setStyleSheet("""
            QPushButton {
                background-color: #5e60ce;
                color: white;
                padding: 6px 15px;
                border-radius: 8px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #4ea8de;
            }
        """)

        self.tb_doanhthunap.setColumnCount(4)
        self.tb_doanhthunap.setHorizontalHeaderLabels(["Mã nạp", "Tên tài khoản", "Thời gian nạp", "Số tiền nạp"])
        self.tb_doanhthunap.setColumnWidth(0,100)
        self.tb_doanhthunap.setColumnWidth(1,100)
        self.tb_doanhthunap.setColumnWidth(2,100)
        self.tb_doanhthunap.setColumnWidth(3,100)
        self.tb_doanhthuthuephong.setColumnCount(6)
        self.tb_doanhthuthuephong.setHorizontalHeaderLabels(["Mã thuê", "Mã máy", "Mã phòng", "Giá tiền","Thời gian thuê","Phương thức"])
        self.tb_doanhthuthuephong.setColumnWidth(0,100)
        self.tb_doanhthuthuephong.setColumnWidth(1,100)
        self.tb_doanhthuthuephong.setColumnWidth(2,100)
        self.tb_doanhthuthuephong.setColumnWidth(3,100)
        self.tb_doanhthuthuephong.setColumnWidth(4,100)
        self.tb_doanhthuthuephong.setColumnWidth(5,100)
        self.tb_doanhthutong.setColumnCount(4)
        self.tb_doanhthutong.setHorizontalHeaderLabels(["Mã", "Thông tin", "Thời gian/Mã phòng", "Số tiền"])
        self.tb_doanhthutong.setColumnWidth(0,100)
        self.tb_doanhthutong.setColumnWidth(1,100)
        self.tb_doanhthutong.setColumnWidth(2,100)
        self.tb_doanhthutong.setColumnWidth(3,100)
        self.tb_doanhthudichvu.setColumnCount(6)
        self.tb_doanhthudichvu.setHorizontalHeaderLabels(["Mã", "Tên tài khoản","Tên dịch vụ","Giá tiền", "Phương thức", "Ngày mua"])
        self.tb_doanhthudichvu.setColumnWidth(0,100)
        self.tb_doanhthudichvu.setColumnWidth(1,100)
        self.tb_doanhthudichvu.setColumnWidth(2,100)
        self.tb_doanhthudichvu.setColumnWidth(3,100)
        self.tb_doanhthudichvu.setColumnWidth(4,100)
        self.tb_doanhthudichvu.setColumnWidth(5,100)
        self.btn_qlmay.clicked.connect(self.go_to_quanlymay)
        self.btn_tbcsvc.clicked.connect(self.go_to_tbcsvc)
        self.btn_dichvu.clicked.connect(self.go_to_dichvu)
        self.btn_nhanvien.clicked.connect(self.go_to_nhanvien)
        self.btn_tk.clicked.connect(self.go_to_taikhoankhach)
        self.btn_tknhanvien.clicked.connect(self.go_to_taikhoannhanvien)
        self.btn_dangxuat.clicked.connect(self.go_to_logout)
        self.btn_trangchu.clicked.connect(self.go_to_trangchu_admin)
        self.btn_qlphong.clicked.connect(self.go_to_quanlyphong)
        self.btn_back.clicked.connect(self.go_to_trangchu_admin)
        self.btn_loc.clicked.connect(self.loc_du_lieu)
        self.menu_nap()
        self.menu_tong()
        self.menu_dichvu()
        self.menu_thue()
        self.loaddata()
        self.resize(1000, 760)
        self.setFixedSize(1000, 760)
        self.ten_tai_khoan = ten_tai_khoan
        self.lb_tentaikhoan.setText(f"Xin chào, {self.ten_tai_khoan}")

            # Lưu query hiện tại sau khi lọc
        self.default_query_nap = "SELECT * FROM lich_su_nap"
        self.default_query_thue = "SELECT * FROM lich_su_thue"
        self.default_query_dichvu = "SELECT * FROM lich_su_dich_vu"
        self.default_query_tong = """
            SELECT ma_nap AS ma, ten_tai_khoan AS thong_tin, thoi_gian_nap AS thoi_gian, so_tien_nap AS so_tien
            FROM lich_su_nap
            UNION ALL
            SELECT ma_ls AS ma, ma_may AS thong_tin, thoi_gian_thue AS thoi_gian, gia_tien AS so_tien
            FROM lich_su_thue
            UNION ALL 
            SELECT ma_lich_su_dich_vu AS ma, ten_tai_khoan AS thong_tin, ngay_mua AS thoi_gian, gia_tien AS so_tien
            FROM lich_su_dich_vu
        """
        # Ban đầu = mặc định
        self.query_nap = self.default_query_nap
        self.query_thue = self.default_query_thue
        self.query_dichvu = self.default_query_dichvu
        self.query_tong = self.default_query_tong


    def menu_nap(self):
        menu = QMenu()
        menu.addAction("Xuất Excel", self.export_excel_nap)
        menu.addAction("Xuất PDF", self.export_pdf_nap)
        self.btn_xuatnap.setMenu(menu)
    def menu_thue(self):
        menu = QMenu()
        menu.addAction("Xuất Excel", self.export_excel_nap)
        menu.addAction("Xuất PDF", self.export_pdf_nap)
        self.btn_xuatthue.setMenu(menu)
    def menu_dichvu(self):
        menu = QMenu()
        menu.addAction("Xuất Excel", self.export_excel_nap)
        menu.addAction("Xuất PDF", self.export_pdf_nap)
        self.btn_xuatdv.setMenu(menu)
    def menu_tong(self):
        menu = QMenu()
        menu.addAction("Xuất Excel", self.export_excel_tong)
        menu.addAction("Xuất PDF", self.export_pdf_tong)
        self.btn_xuattong.setMenu(menu)
    def setup_actions(self):
        self.btn_loc.clicked.connect(self.loc_du_lieu)
        self.menu_nap()
        self.menu_tong()
        self.loaddata() 
    def loc_du_lieu(self):
        try:
            conn = mysql.connector.connect(
                host="localhost", user="root", password="", database="qly_quan_net"
            )
            cursor = conn.cursor()

            chedo = self.cbb_thongke.currentText()  # Lấy chế độ lọc từ combobox
            ngay = self.date_ngay.date().toString("yyyy-MM-dd")
            thang = self.date_thang.date().toString("yyyy-MM")

            if chedo == "Theo ngày":
                self.query_nap = f"SELECT * FROM lich_su_nap WHERE DATE(thoi_gian_nap) = '{ngay}'"
                self.query_thue = f"SELECT * FROM lich_su_thue WHERE DATE(thoi_gian_thue) = '{ngay}'"
                self.query_dichvu = f"SELECT * FROM lich_su_dich_vu WHERE DATE(ngay_mua) = '{ngay}'"
                self.query_tong = f"""
                    SELECT ma_nap AS ma, ten_tai_khoan AS thong_tin, thoi_gian_nap AS thoi_gian, so_tien_nap AS so_tien
                    FROM lich_su_nap WHERE DATE(thoi_gian_nap) = '{ngay}'
                    UNION ALL
                    SELECT ma_ls AS ma, ma_may AS thong_tin, thoi_gian_thue AS thoi_gian, gia_tien AS so_tien
                    FROM lich_su_thue WHERE DATE(thoi_gian_thue) = '{ngay}'
                    UNION ALL
                    SELECT ma_lich_su_dich_vu AS ma, ten_tai_khoan AS thong_tin, ngay_mua AS thoi_gian, gia_tien AS so_tien
                    FROM lich_su_dich_vu WHERE DATE(ngay_mua) = '{ngay}'
                """

            elif chedo == "Theo tháng":
                self.query_nap = f"SELECT * FROM lich_su_nap WHERE DATE_FORMAT(thoi_gian_nap, '%Y-%m') = '{thang}'"
                self.query_thue = f"SELECT * FROM lich_su_thue WHERE DATE_FORMAT(thoi_gian_thue, '%Y-%m') = '{thang}'"
                self.query_dichvu = f"SELECT * FROM lich_su_dich_vu WHERE DATE_FORMAT(ngay_mua, '%Y-%m') = '{thang}'"
                self.query_tong = f"""
                    SELECT ma_nap AS ma, ten_tai_khoan AS thong_tin, thoi_gian_nap AS thoi_gian, so_tien_nap AS so_tien
                    FROM lich_su_nap WHERE DATE_FORMAT(thoi_gian_nap, '%Y-%m') = '{thang}'
                    UNION ALL
                    SELECT ma_ls AS ma, ma_may AS thong_tin, thoi_gian_thue AS thoi_gian, gia_tien AS so_tien
                    FROM lich_su_thue WHERE DATE_FORMAT(thoi_gian_thue, '%Y-%m') = '{thang}'
                    UNION ALL
                    SELECT ma_lich_su_dich_vu AS ma, ten_tai_khoan AS thong_tin, ngay_mua AS thoi_gian, gia_tien AS so_tien
                    FROM lich_su_dich_vu WHERE DATE_FORMAT(ngay_mua, '%Y-%m') = '{thang}'
                """

            elif chedo == "Tổng":  # Reset về mặc định
                self.query_nap = self.default_query_nap
                self.query_thue = self.default_query_thue
                self.query_dichvu = self.default_query_dichvu
                self.query_tong = self.default_query_tong

            # load lại dữ liệu
            self.loaddata()

            cursor.close()
            conn.close()

        except mysql.connector.Error as e:
            QMessageBox.critical(self, "Lỗi MySQL", f"Lỗi khi lọc dữ liệu: {e}")

    def loaddata(self):
        try:
            conn = mysql.connector.connect(host="localhost", user="root", password="", database="qly_quan_net")
            cursor = conn.cursor()

            #  bảng nạp 
            cursor.execute(getattr(self, "query_nap", "SELECT * FROM lich_su_nap"))
            data = cursor.fetchall()
            self.tb_doanhthunap.setRowCount(len(data))
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    self.tb_doanhthunap.setItem(r, c, QTableWidgetItem(str(val)))

            #  bảng thuê 
            cursor.execute(getattr(self, "query_thue", "SELECT * FROM lich_su_thue"))
            data = cursor.fetchall()
            self.tb_doanhthuthuephong.setRowCount(len(data))
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    self.tb_doanhthuthuephong.setItem(r, c, QTableWidgetItem(str(val)))
            #  bảng dichvu
            cursor.execute(getattr(self, "query_dichvu", "SELECT * FROM lich_su_dich_vu"))
            data = cursor.fetchall()
            self.tb_doanhthudichvu.setRowCount(len(data))
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    self.tb_doanhthudichvu.setItem(r, c, QTableWidgetItem(str(val)))
            #  bảng tổng 
            cursor.execute(getattr(self, "query_tong", """
                SELECT ma_nap AS ma, ten_tai_khoan AS thong_tin, thoi_gian_nap AS thoi_gian, so_tien_nap AS so_tien
                FROM lich_su_nap
                UNION ALL
                SELECT ma_ls AS ma, ma_may AS thong_tin, thoi_gian_thue AS thoi_gian, gia_tien AS so_tien
                FROM lich_su_thue
                UNION ALL 
                SELECT ma_lich_su_dich_vu AS ma, ten_tai_khoan AS thong_tin, ngay_mua AS thoi_gian, gia_tien AS so_tien
                FROM lich_su_dich_vu
            """))
            data = cursor.fetchall()
            self.tb_doanhthutong.setRowCount(len(data))
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    self.tb_doanhthutong.setItem(r, c, QTableWidgetItem(str(val)))

            cursor.close()
            conn.close()

        except mysql.connector.Error as e:
            print(f"Lỗi MySQL: {e}")

    # ==== Xuất Excel chung ====
    def export_excel(self, query, title):
        try:
            conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="",
                database="qly_quan_net"
            )
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]

            file_path, _ = QFileDialog.getSaveFileName(
                self, f"Lưu {title} Excel", "", "Excel Files (*.xlsx)"
            )
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = title

            # Header
            for col_num, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            # Data
            for row in rows:
                ws.append(row)

            # Autofit column width
            for col_num, col_name in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col_letter])
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(file_path)
            QMessageBox.information(self, "Thành công", f"Xuất {title} ra Excel: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể xuất Excel: {e}")

    # ==== Xuất PDF chung ====
    def export_pdf(self, query, title):
        try:
            conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="",
                database="qly_quan_net"
            )
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]

            file_path, _ = QFileDialog.getSaveFileName(
                self, f"Lưu {title} PDF", "", "PDF Files (*.pdf)"
            )
            if not file_path:
                return

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            elements.append(Paragraph(title.upper(), styles["Title"]))
            elements.append(Spacer(1, 12))

            # Data table
            data = [columns] + rows
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ]))
            elements.append(table)

            doc.build(elements)
            QMessageBox.information(self, "Thành công", f"Xuất {title} ra PDF: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể xuất PDF: {e}")

    # ==== Xuất Excel ====
    def export_excel_nap(self):
        query = self.query_nap or self.default_query_nap
        self.export_excel(query, "DoanhThuNap")

    def export_excel_tong(self):
        query = self.query_tong or self.default_query_tong
        self.export_excel(query, "DoanhThuTong")
    def export_excel_thue(self):
        query = self.query_thue or self.default_query_thue
        self.export_excel(query, "DoanhThuThuea")
    def export_excel_dichvu(self):
        query = self.query_dichvu or self.default_query_dichvu
        self.export_excel(query, "DoanhThuDichVu")

    # ==== Xuất PDF ====
    def export_pdf_nap(self):
        query = self.query_nap or self.default_query_nap
        self.export_pdf(query, "Doanh Thu Nạp")

    def export_pdf_tong(self):
        query = self.query_tong or self.default_query_tong
        self.export_pdf(query, "Doanh Thu Tổng")
    def export_pdf_thue(self):
        query = self.query_thue or self.default_query_thue
        self.export_pdf(query, "Doanh Thu Thuê Phòng")
    def export_pdf_dichvu(self):
        query = self.query_dichvu or self.default_query_dichvu
        self.export_pdf(query, "Doanh Thu Dịch Vụ")

    def go_to_quanlyphong(self):
        from quanlyphong import quanlyphong
        self.quanlyphong_form = quanlyphong(self.ten_tai_khoan)
        self.quanlyphong_form.show()
        self.hide()
    def go_to_quanlymay(self):
        from quanlymay import quanlymay
        self.quanlymay_form = quanlymay(self.ten_tai_khoan)
        self.quanlymay_form.show()
        self.hide()
    def go_to_tbcsvc(self):
        from thietbi_cosovatchat import thietbi_cosovatchat
        self.thietbi_cosovatchat_form = thietbi_cosovatchat(self.ten_tai_khoan)
        self.thietbi_cosovatchat_form.show()
        self.hide()
    def go_to_dichvu(self):
        from dich_vu import dich_vu
        self.dich_vu_form = dich_vu(self.ten_tai_khoan)
        self.dich_vu_form.show()
        self.hide()
    def go_to_nhanvien(self):
        from nhan_vien import nhan_vien
        self.nhan_vien_form = nhan_vien(self.ten_tai_khoan)
        self.nhan_vien_form.show()
        self.hide()
    def go_to_taikhoankhach(self):
        from taikhoan_khach import taikhoan_khach
        self.taikhoan_khach_form = taikhoan_khach(self.ten_tai_khoan)
        self.taikhoan_khach_form.show()
        self.hide()
    def go_to_taikhoannhanvien(self):
        from taikhoan_nhanvien import taikhoan_nhanvien
        self.taikhoan_nhanvien_form = taikhoan_nhanvien(self.ten_tai_khoan)
        self.taikhoan_nhanvien_form.show()
        self.hide()
    def go_to_logout(self):
        from dangnhap_admin import dangnhap_admin
        self.dangnhap_admin_form = dangnhap_admin()
        self.dangnhap_admin_form.show()
        self.hide()
    def go_to_trangchu_admin(self):
        from trangchu_admin import trangchu_admin
        self.trangchu_admin_form = trangchu_admin(self.ten_tai_khoan)
        self.trangchu_admin_form.show()
        self.hide()
    def go_to_phong(self):
        from quanlyphong import quanlyphong
        self.quanlyphong_form = quanlyphong(self.ten_tai_khoan)
        self.quanlyphong_form.show()
        self.hide()
    def go_to_bao_cao(self):
        from baocao_thongke import baocao_thongke
        self.baocao_thongke_form = baocao_thongke(self.ten_tai_khoan)
        self.baocao_thongke_form.show()
        self.hide()
# app = QApplication(sys.argv)
# widget = QtWidgets.QStackedWidget()
# baocao_thongke_form = baocao_thongke()
# widget.addWidget(baocao_thongke_form)
# widget.setCurrentWidget(baocao_thongke_form)

# widget.resize(1000, 760)
# widget.show()
# app.exec()