import sys
import unittest
from unittest.mock import patch, MagicMock
from PyQt6.QtWidgets import QApplication
from index.nhan_vien import nhan_vien
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class CustomTestResult(unittest.TestResult):
    def __init__(self, stream=None, verbosity=1):
        super().__init__()
        self.stream = stream
        self.verbosity = verbosity
        self.test_results = []
        
    def addSuccess(self, test):
        super().addSuccess(test)
        self._record_result(test, "PASS", "")

    def addError(self, test, err):
        super().addError(test, err)
        self._record_result(test, "ERROR", str(err[1]))

    def addFailure(self, test, err):
        super().addFailure(test, err)
        self._record_result(test, "FAIL", str(err[1]))

    def _record_result(self, test, result, details=""):
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, "_input_data", "")
        expected_data = getattr(test, "_expected_data", "")
        self.test_results.append({
            "name": test_name,
            "result": result,
            "details": details,
            "input_data": input_data,
            "expected_data": expected_data,
        })

class CustomTestRunner:
    def __init__(self, stream=None, verbosity=1):
        self.stream = stream if stream else sys.stdout
        self.verbosity = verbosity
        
    def run(self, test):
        result = CustomTestResult(self.stream, self.verbosity)
        test(result)
        self.print_results(result)
        self.export_to_excel(result)
        return result
    
    def print_results(self, result):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("\n===== KẾT QUẢ TEST =====")
        print(f"Thời gian chạy: {current_time}\n")
        for idx, r in enumerate(result.test_results, 1):
            print(f"--- Test {idx} ---")
            print(f"Test Case        : {r['name']}")
            print(f"Kết quả          : {r['result']}")
            print(f"Dữ liệu đầu vào  : {r['input_data']}")
            print(f"Kết quả mong đợi : {r['expected_data']}")
            if r["details"]:
                print(f"Chi tiết         : {r['details']}")
            print("-" * 40)

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Thông tin chung
        ws["A1"] = "Function Code"; ws["B1"] = "nhan_vien"
        ws["A2"] = "Function Name"; ws["B2"] = "Thêm nhân viên"
        ws["A3"] = "Created By"; ws["B3"] = "tester"
        ws["A4"] = "Test requirement"; ws["B4"] = "Kiểm tra chức năng thêm nhân viên với các trường hợp đầu vào"

        # Tạo header ở row 6
        ws.append([])  # row 5 trống
        headers = ["STT", "Test Case", "Dữ liệu đầu vào", "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái"]
        ws.append(headers)

        # Định dạng header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=6, column=col)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        # Ghi dữ liệu test
        for idx, r in enumerate(result.test_results, 1):
            ws.append([
                idx,
                r["name"],
                r["input_data"],
                r["expected_data"],
                r["details"] if r["details"] else r["expected_data"],
                r["result"],
            ])

        # Auto-fit column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        filename = f"nhanvien.xlsx"
        wb.save(filename)
        print(f"Đã xuất kết quả ra file {filename}")

app = QApplication(sys.argv)

class TestNhanVien(unittest.TestCase):
    def setUp(self):
        """Thiết lập trước mỗi test case"""
        with patch('mysql.connector.connect') as mock_connect:
            mock_conn = MagicMock()
            mock_cursor = MagicMock()
            mock_connect.return_value = mock_conn
            mock_conn.cursor.return_value = mock_cursor
            
            self.nv = nhan_vien("admin_test")
            self.nv.conn = mock_conn
            self.nv.cursor = mock_cursor

    def tearDown(self):
        """Dọn dẹp sau mỗi test case"""
        if hasattr(self.nv, 'conn'):
            self.nv.conn.close()

    @patch("index.nhan_vien.QMessageBox")
    def test_them_thieu_thong_tin(self, mock_msgbox):
        """Test thiếu thông tin"""
        self._input_data = "Tên='', SĐT='', Email=''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.nv.txt_ten.setText("")
        self.nv.txt_sodienthoai.setText("")
        self.nv.txt_mail.setText("")
        self.nv.txt_diachi.setText("Hà Nội")
        self.nv.cbb_maphong.setCurrentText("P001")
        self.nv.cbb_chucvu.setCurrentText("Nhân viên")
        
        self.nv.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.nhan_vien.QMessageBox")
    def test_them_sdt_sai_format(self, mock_msgbox):
        """Test số điện thoại sai format"""
        self._input_data = "Tên='Nguyễn Văn A', SĐT='123456789'"
        self._expected_data = "Hiển thị lỗi format số điện thoại"
        
        self.nv.txt_ten.setText("Nguyễn Văn A")
        self.nv.txt_sodienthoai.setText("123456789")
        self.nv.txt_mail.setText("test@gmail.com")
        self.nv.txt_diachi.setText("Hà Nội")
        self.nv.cbb_maphong.setCurrentText("P001")
        self.nv.cbb_chucvu.setCurrentText("Nhân viên")
        
        self.nv.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.nhan_vien.QMessageBox")
    def test_them_ten_co_so(self, mock_msgbox):
        """Test tên có chứa số"""
        self._input_data = "Tên='Nguyễn Văn A123', SĐT='0123456789'"
        self._expected_data = "Hiển thị lỗi format tên"
        
        self.nv.txt_ten.setText("Nguyễn Văn A123")
        self.nv.txt_sodienthoai.setText("0123456789")
        self.nv.txt_mail.setText("test@gmail.com")
        self.nv.txt_diachi.setText("Hà Nội")
        self.nv.cbb_maphong.setCurrentText("P001")
        self.nv.cbb_chucvu.setCurrentText("Nhân viên")
        
        self.nv.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.nhan_vien.QMessageBox")
    def test_them_email_sai_format(self, mock_msgbox):
        """Test email sai format"""
        self._input_data = "Tên='Nguyễn Văn A', Email='test@yahoo.com'"
        self._expected_data = "Hiển thị lỗi email phải có đuôi @gmail.com"
        
        self.nv.txt_ten.setText("Nguyễn Văn A")
        self.nv.txt_sodienthoai.setText("0123456789")
        self.nv.txt_mail.setText("test@yahoo.com")
        self.nv.txt_diachi.setText("Hà Nội")
        self.nv.cbb_maphong.setCurrentText("P001")
        self.nv.cbb_chucvu.setCurrentText("Nhân viên")
        
        self.nv.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.nhan_vien.QMessageBox")
    def test_them_email_da_ton_tai(self, mock_msgbox):
        """Test email đã tồn tại"""
        self._input_data = "Tên='Nguyễn Văn A', Email='existing@gmail.com'"
        self._expected_data = "Hiển thị lỗi email đã tồn tại"
        
        # Mock database trả về có email trùng
        self.nv.cursor.fetchone.return_value = ("NV001",)
        
        self.nv.txt_ten.setText("Nguyễn Văn A")
        self.nv.txt_sodienthoai.setText("0123456789")
        self.nv.txt_mail.setText("existing@gmail.com")
        self.nv.txt_diachi.setText("Hà Nội")
        self.nv.cbb_maphong.setCurrentText("P001")
        self.nv.cbb_chucvu.setCurrentText("Nhân viên")
        
        self.nv.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.nhan_vien.QMessageBox")
    def test_them_hop_le(self, mock_msgbox):
        """Test thêm nhân viên hợp lệ"""
        self._input_data = "Tên='Nguyễn Văn A', SĐT='0123456789', Email='valid@gmail.com'"
        self._expected_data = "Thêm nhân viên thành công"
        
        # Mock database không có dữ liệu trùng
        self.nv.cursor.fetchone.return_value = None
        
        self.nv.txt_ten.setText("Nguyễn Văn A")
        self.nv.txt_sodienthoai.setText("0123456789")
        self.nv.txt_mail.setText("valid@gmail.com")
        self.nv.txt_diachi.setText("Hà Nội")
        self.nv.cbb_maphong.setCurrentText("P001")
        self.nv.cbb_chucvu.setCurrentText("Nhân viên")
        
        # Mock auto generate mã nhân viên
        with patch.object(self.nv, 'tao_tu_dong_manhanvien', return_value='NV001'):
            self.nv.them()
        
        # Kiểm tra có gọi execute (thêm vào DB)
        self.assertTrue(self.nv.cursor.execute.called)
if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestNhanVien)
    runner = CustomTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Tính toán thống kê
    total_tests = len(result.test_results)
    passed = len([r for r in result.test_results if r['result'] == 'PASS'])
    failed = len([r for r in result.test_results if r['result'] == 'FAIL'])
    errors = len([r for r in result.test_results if r['result'] == 'ERROR'])
    
    print(f"\n===== THỐNG KÊ =====")
    print(f"Tổng số test: {total_tests}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")
    print(f"Errors: {errors}")
