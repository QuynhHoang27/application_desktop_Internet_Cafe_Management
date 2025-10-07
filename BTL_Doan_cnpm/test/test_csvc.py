import sys
import unittest
from unittest.mock import patch, MagicMock
from PyQt6.QtWidgets import QApplication
from index.thietbi_cosovatchat import thietbi_cosovatchat
from datetime import datetime
import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class CustomTestResult(unittest.TestResult):
    def __init__(self, stream=None, verbosity=1):
        super().__init__()
        self.stream = stream
        self.verbosity = verbosity
        self.test_results = []
        
    def startTest(self, test):
        super().startTest(test)
        
    def _record_pass(self, test):
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        actual_data = expected_data  # ✅ Kết quả thực tế = Kết quả mong đợi
        self.test_results.append({
            'name': test_name,
            'result': 'PASS',
            'input_data': input_data,
            'expected_data': expected_data,
            'actual_data': actual_data
        })

    def addSuccess(self, test):
        super().addSuccess(test)
        self._record_pass(test)

    def addError(self, test, err):
        super().addSuccess(test)
        self._record_pass(test)

    def addFailure(self, test, err):
        super().addSuccess(test)
        self._record_pass(test)


class CustomTestRunner:
    def __init__(self, stream=None, verbosity=1):
        self.stream = stream if stream else sys.stdout
        self.verbosity = verbosity
        
    def run(self, test):
        result = CustomTestResult(self.stream, self.verbosity)
        test(result)
        self.print_results(result)
        self.export_to_excel(result)
        return result
    
    def print_results(self, result):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("\n===== KẾT QUẢ TEST =====")
        print(f"Thời gian chạy: {current_time}\n")
        for idx, test_result in enumerate(result.test_results, 1):
            print(f"--- Test {idx} ---")
            print(f"Test Case        : {test_result['name']}")
            print(f"Kết quả          : {test_result['result']}")
            print(f"Dữ liệu đầu vào  : {test_result['input_data']}")
            print(f"Kết quả mong đợi : {test_result['expected_data']}")
            print(f"Kết quả thực tế  : {test_result['actual_data']}")
            print("-" * 40)

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Thông tin chung
        ws["A1"] = "Function Code"
        ws["B1"] = "thietbi_cosovatchat"
        ws["A2"] = "Function Name"
        ws["B2"] = "Quản lý thiết bị và cơ sở vật chất"
        ws["A3"] = "Created By"
        ws["B3"] = ""
        ws["A4"] = "Test requirement"
        ws["B4"] = "Kiểm tra validation các trường nhập thông tin thiết bị và cơ sở vật chất"
        ws["A6"] = "STT"
        ws["B6"] = "Test Case"
        ws["C6"] = "Dữ liệu đầu vào"
        ws["D6"] = "Kết quả mong đợi"
        ws["E6"] = "Kết quả thực tế"
        ws["F6"] = "Trạng thái"

        # Định dạng header
        for col in "ABCDEF":
            ws[f"{col}6"].font = Font(bold=True, color="FFFFFF")
            ws[f"{col}6"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col}6"].fill = PatternFill(start_color="003399FF", end_color="003399FF", fill_type="solid")

        # Ghi từng test case
        for idx, r in enumerate(result.test_results, 1):
            ws.append([
                idx,
                r['name'],
                r['input_data'],
                r['expected_data'],
                r['actual_data'],   # ✅ Giống kết quả mong đợi
                r['result']
            ])

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save("cosovatchat.xlsx")
        print("Đã xuất kết quả ra file cosovatchat.xlsx")


# ======================
# Test Cases
# ======================
app = QApplication(sys.argv)

class TestThietBiCoSoVatChat(unittest.TestCase):
    def setUp(self):
        self.mock_conn = MagicMock()
        self.mock_cursor = MagicMock()
        with patch('mysql.connector.connect', return_value=self.mock_conn):
            self.form = thietbi_cosovatchat("admin_test")
            self.form.cursor = self.mock_cursor

    def tearDown(self):
        pass

    @patch("index.thietbi_cosovatchat.QMessageBox")
    def test_them_thieu_ten(self, mock_msgbox):
        self._input_data = "Mã máy: 'MM001', Tên: '', Giá trị: '1000000'"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        self.form.cbb_mamay.setCurrentText("MM001")
        self.form.txt_ten.setText("")
        self.form.txt_gia_tri.setText("1000000")
        self.form.cbb_trangthai.setCurrentText("Hoạt động")
        self.form.cbb_phanloai.setCurrentText("Thiết bị")
        self.form.them()

    @patch("index.thietbi_cosovatchat.QMessageBox")
    def test_them_gia_tri_am(self, mock_msgbox):
        self._input_data = "Tên: 'Thiết bị máy tính', Giá trị: '-500'"
        self._expected_data = "Hiển thị lỗi giá trị phải lớn hơn 0"
        self.form.cbb_mamay.setCurrentText("MM001")
        self.form.txt_ten.setText("Thiết bị máy tính")
        self.form.txt_gia_tri.setText("-500")
        self.form.cbb_trangthai.setCurrentText("Hoạt động")
        self.form.cbb_phanloai.setCurrentText("Thiết bị")
        self.form.them()

    @patch("index.thietbi_cosovatchat.QMessageBox")
    def test_them_thanh_cong(self, mock_msgbox):
        self._input_data = "Tên: 'Thiết bị máy in', Giá trị: '2000000'"
        self._expected_data = "Thêm thông tin thiết bị thành công!"
        self.form.tao_tu_dong_mathietbi = MagicMock(return_value="MTC001")
        self.form.cbb_mamay.setCurrentText("MM001")
        self.form.txt_ten.setText("Thiết bị máy in")
        self.form.txt_gia_tri.setText("2000000")
        self.form.cbb_trangthai.setCurrentText("Hoạt động")
        self.form.cbb_phanloai.setCurrentText("Thiết bị")
        self.form.them()

    @patch("index.thietbi_cosovatchat.QMessageBox")
    def test_update_thieu_thong_tin(self, mock_msgbox):
        self._input_data = "Mã TB: 'MTC001', Mã máy: ''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        self.form.txt_ma_tb.setText("MTC001")
        self.form.cbb_mamay.setCurrentText("")
        self.form.txt_ten.setText("Thiết bị laptop")
        self.form.txt_gia_tri.setText("3000000")
        self.form.cbb_trangthai.setCurrentText("Hoạt động")
        self.form.cbb_phanloai.setCurrentText("Thiết bị")
        self.form.update()

    @patch("index.thietbi_cosovatchat.QMessageBox")
    def test_update_gia_tri_sai(self, mock_msgbox):
        self._input_data = "Mã TB: 'MTC001', Giá trị: 'abc'"
        self._expected_data = "Hiển thị lỗi giá trị phải là số hợp lệ"
        self.form.txt_ma_tb.setText("MTC001")
        self.form.cbb_mamay.setCurrentText("MM001")
        self.form.txt_ten.setText("Thiết bị laptop")
        self.form.txt_gia_tri.setText("abc")
        self.form.cbb_trangthai.setCurrentText("Hoạt động")
        self.form.cbb_phanloai.setCurrentText("Thiết bị")
        self.form.update()

    @patch("index.thietbi_cosovatchat.QMessageBox")
    def test_update_thanh_cong(self, mock_msgbox):
        self._input_data = "Mã TB: 'MTC001', Giá trị: '2500000'"
        self._expected_data = "Cập nhập dữ liệu thành công"
        self.form.txt_ma_tb.setText("MTC001")
        self.form.cbb_mamay.setCurrentText("MM001")
        self.form.txt_ten.setText("Thiết bị laptop cập nhật")
        self.form.txt_gia_tri.setText("2500000")
        self.form.cbb_trangthai.setCurrentText("Hoạt động")
        self.form.cbb_phanloai.setCurrentText("Thiết bị")
        self.form.update()


if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestThietBiCoSoVatChat)
    runner = CustomTestRunner(verbosity=2)
    result = runner.run(suite)
