import sys
import unittest
from unittest.mock import patch, MagicMock
from PyQt6.QtWidgets import QApplication
from index.dich_vu import dich_vu
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
class CustomTestResult(unittest.TestResult):
    def __init__(self, stream=None, verbosity=1):
        super().__init__()
        self.stream = stream
        self.verbosity = verbosity
        self.test_results = []
        
    def addSuccess(self, test):
        super().addSuccess(test)
        self._record_result(test, "PASS", "")

    def addError(self, test, err):
        super().addError(test, err)
        self._record_result(test, "ERROR", str(err[1]))

    def addFailure(self, test, err):
        super().addFailure(test, err)
        self._record_result(test, "FAIL", str(err[1]))

    def _record_result(self, test, result, details=""):
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, "_input_data", "")
        expected_data = getattr(test, "_expected_data", "")
        self.test_results.append({
            "name": test_name,
            "result": result,
            "details": details,
            "input_data": input_data,
            "expected_data": expected_data,
        })



class CustomTestRunner:
    def __init__(self, stream=None, verbosity=1):
        self.stream = stream if stream else sys.stdout
        self.verbosity = verbosity
        
    def run(self, test):
        result = CustomTestResult(self.stream, self.verbosity)
        test(result)
        self.print_results(result)
        self.export_to_excel(result)
        return result
    
    def print_results(self, result):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("\n===== KẾT QUẢ TEST =====")
        print(f"Thời gian chạy: {current_time}\n")
        for idx, r in enumerate(result.test_results, 1):
            print(f"--- Test {idx} ---")
            print(f"Test Case        : {r['name']}")
            print(f"Kết quả          : {r['result']}")
            print(f"Dữ liệu đầu vào  : {r['input_data']}")
            print(f"Kết quả mong đợi : {r['expected_data']}")
            if r["details"]:
                print(f"Chi tiết         : {r['details']}")
            print("-" * 40)

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Thông tin chung
        ws["A1"] = "Function Code"; ws["B1"] = "dich_vu"
        ws["A2"] = "Function Name"; ws["B2"] = "Thêm dịch vụ"
        ws["A3"] = "Created By"; ws["B3"] = "tester"
        ws["A4"] = "Test requirement"; ws["B4"] = "Kiểm tra chức năng thêm dịch vụ với các trường hợp đầu vào"

        # Tạo header ở row 6
        ws.append([])  # row 5 trống
        headers = ["STT", "Test Case", "Dữ liệu đầu vào", "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái"]
        ws.append(headers)

        # Định dạng header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=6, column=col)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        # Ghi dữ liệu test
        for idx, r in enumerate(result.test_results, 1):
            ws.append([
                idx,
                r["name"],
                r["input_data"],
                r["expected_data"],
                r["details"] if r["details"] else r["expected_data"],
                r["result"],
            ])

        # Auto-fit column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Giới hạn width tối đa 50
            ws.column_dimensions[col_letter].width = adjusted_width

        filename = f"dichvu.xlsx"
        wb.save(filename)
        print(f"Đã xuất kết quả ra file {filename}")


app = QApplication(sys.argv)

class TestDichVu(unittest.TestCase):
    def setUp(self):
        """Thiết lập trước mỗi test case"""
        with patch('mysql.connector.connect') as mock_connect:
            mock_conn = MagicMock()
            mock_cursor = MagicMock()
            mock_connect.return_value = mock_conn
            mock_conn.cursor.return_value = mock_cursor
            
            self.dv = dich_vu("admin_test")
            self.dv.conn = mock_conn
            self.dv.cursor = mock_cursor

    def tearDown(self):
        """Dọn dẹp sau mỗi test case"""
        if hasattr(self.dv, 'conn'):
            self.dv.conn.close()

    @patch("index.dich_vu.QMessageBox")
    def test_them_thieu_ten(self, mock_msgbox):
        """Test thiếu tên dịch vụ"""
        self._input_data = "Tên='', Giá='15000'"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.dv.txt_ten.setText("")
        self.dv.txt_gia_tri.setText("15000")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Đảm bảo method được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_thieu_gia(self, mock_msgbox):
        """Test thiếu giá trị"""
        self._input_data = "Tên='Đồ ăn: Mì tôm', Giá=''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.dv.txt_ten.setText("Đồ ăn: Mì tôm")
        self.dv.txt_gia_tri.setText("")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Đảm bảo method được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_gia_tri_khong_phai_so(self, mock_msgbox):
        """Test giá trị không phải số"""
        self._input_data = "Tên='Đồ ăn: Mì tôm', Giá='abc'"
        self._expected_data = "Hiển thị lỗi giá trị phải là số nguyên không âm"
        
        self.dv.txt_ten.setText("Đồ ăn: Mì tôm")
        self.dv.txt_gia_tri.setText("abc")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_gia_tri_am(self, mock_msgbox):
        """Test giá trị âm"""
        self._input_data = "Tên='Đồ ăn: Mì tôm', Giá='-1000'"
        self._expected_data = "Hiển thị lỗi giá trị phải lớn hơn 0"
        
        self.dv.txt_ten.setText("Đồ ăn: Mì tôm")
        self.dv.txt_gia_tri.setText("-1000")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_gia_tri_bang_khong(self, mock_msgbox):
        """Test giá trị bằng 0"""
        self._input_data = "Tên='Đồ ăn: Mì tôm', Giá='0'"
        self._expected_data = "Hiển thị lỗi giá trị phải lớn hơn 0"
        
        self.dv.txt_ten.setText("Đồ ăn: Mì tôm")
        self.dv.txt_gia_tri.setText("0")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_sai_dinh_dang_do_an(self, mock_msgbox):
        """Test sai định dạng tên đồ ăn"""
        self._input_data = "Tên='Mì tôm', Loại='Đồ ăn'"
        self._expected_data = "Hiển thị lỗi sai định dạng tên dịch vụ"
        
        self.dv.txt_ten.setText("Mì tôm")
        self.dv.txt_gia_tri.setText("15000")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_sai_dinh_dang_nuoc_uong(self, mock_msgbox):
        """Test sai định dạng tên nước uống"""
        self._input_data = "Tên='Coca Cola', Loại='Nước uống'"
        self._expected_data = "Hiển thị lỗi sai định dạng tên dịch vụ"
        
        self.dv.txt_ten.setText("Coca Cola")
        self.dv.txt_gia_tri.setText("12000")
        self.dv.cbb_phanloai.setCurrentText("Nước uống")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_sai_dinh_dang_khac(self, mock_msgbox):
        """Test sai định dạng tên loại khác"""
        self._input_data = "Tên='Thuê tai nghe', Loại='Khác'"
        self._expected_data = "Hiển thị lỗi sai định dạng tên dịch vụ"
        
        self.dv.txt_ten.setText("Thuê tai nghe")
        self.dv.txt_gia_tri.setText("5000")
        self.dv.cbb_phanloai.setCurrentText("Khác")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_ten_da_ton_tai(self, mock_msgbox):
        """Test tên dịch vụ đã tồn tại"""
        self._input_data = "Tên='Đồ ăn: Mì tôm', Giá='15000' (đã tồn tại)"
        self._expected_data = "Hiển thị lỗi tên dịch vụ đã tồn tại"
        
        # Mock database trả về có dữ liệu trùng
        self.dv.cursor.fetchone.return_value = ("DV001",)
        
        self.dv.txt_ten.setText("Đồ ăn: Mì tôm")
        self.dv.txt_gia_tri.setText("15000")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        self.dv.them()
        
        # Kiểm tra MessageBox được gọi
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_hop_le_do_an(self, mock_msgbox):
        """Test thêm đồ ăn hợp lệ"""
        self._input_data = "Tên='Đồ ăn: Cơm chiên', Giá='25000'"
        self._expected_data = "Thêm dịch vụ thành công"
        
        # Mock database không có dữ liệu trùng
        self.dv.cursor.fetchone.return_value = None
        
        self.dv.txt_ten.setText("Đồ ăn: Cơm chiên")
        self.dv.txt_gia_tri.setText("25000")
        self.dv.cbb_phanloai.setCurrentText("Đồ ăn")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        # Mock auto generate mã dịch vụ
        with patch.object(self.dv, 'tao_tu_dong_madichvu', return_value='DV001'):
            self.dv.them()
        
        # Kiểm tra có gọi execute (thêm vào DB)
        self.assertTrue(self.dv.cursor.execute.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_hop_le_nuoc_uong(self, mock_msgbox):
        """Test thêm nước uống hợp lệ"""
        self._input_data = "Tên='Nước uống: Pepsi', Giá='12000'"
        self._expected_data = "Thêm dịch vụ thành công"
        
        # Mock database không có dữ liệu trùng
        self.dv.cursor.fetchone.return_value = None
        
        self.dv.txt_ten.setText("Nước uống: Pepsi")
        self.dv.txt_gia_tri.setText("12000")
        self.dv.cbb_phanloai.setCurrentText("Nước uống")
        self.dv.cbb_trangthai.setCurrentText("Còn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
     
        with patch.object(self.dv, 'tao_tu_dong_madichvu', return_value='DV002'):
            self.dv.them()
        
        # Kiểm tra có gọi execute (thêm vào DB)
        self.assertTrue(self.dv.cursor.execute.called)

    @patch("index.dich_vu.QMessageBox")
    def test_them_hop_le_khac(self, mock_msgbox):
        """Test thêm dịch vụ loại khác hợp lệ"""
        self._input_data = "Tên='Khác: In tài liệu', Giá='2000'"
        self._expected_data = "Thêm dịch vụ thành công"
        
        # Mock database không có dữ liệu trùng
        self.dv.cursor.fetchone.return_value = None
        
        self.dv.txt_ten.setText("Khác: In tài liệu")
        self.dv.txt_gia_tri.setText("2000")
        self.dv.cbb_phanloai.setCurrentText("Khác")
        self.dv.cbb_trangthai.setCurrentText("Có sẵn")
        self.dv.cbb_maphong.setCurrentText("P001")
        
        # Mock auto generate mã dịch vụ
        with patch.object(self.dv, 'tao_tu_dong_madichvu', return_value='DV003'):
            self.dv.them()
        
        # Kiểm tra có gọi execute (thêm vào DB)
        self.assertTrue(self.dv.cursor.execute.called)

if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestDichVu)
    runner = CustomTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Tính toán thống kê
    total_tests = len(result.test_results)
    passed = len([r for r in result.test_results if r['result'] == 'PASS'])
    failed = len([r for r in result.test_results if r['result'] == 'FAIL'])
    errors = len([r for r in result.test_results if r['result'] == 'ERROR'])
    
    print(f"\n===== THỐNG KÊ =====")
    print(f"Tổng số test: {total_tests}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")
    print(f"Errors: {errors}")
  