#python -m test.dangnhap_admin_test
import sys
import unittest
from unittest.mock import patch
from PyQt6.QtWidgets import QApplication
from index.dangnhap_admin import dangnhap_admin
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class CustomTestResult(unittest.TestResult):
    def __init__(self, stream=None, verbosity=1):
        super().__init__()
        self.stream = stream
        self.verbosity = verbosity
        self.test_results = []
        
    def startTest(self, test):
        super().startTest(test)
        
    def addSuccess(self, test):
        super().addSuccess(test)
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        self.test_results.append({
            'name': test_name,
            'result': 'PASS',
            'details': '',
            'input_data': input_data,
            'expected_data': expected_data
        })
        
    def addError(self, test, err):
        super().addError(test, err)
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        self.test_results.append({
            'name': test_name,
            'result': 'ERROR',
            'details': str(err[1]),
            'input_data': input_data,
            'expected_data': expected_data
        })
        
    def addFailure(self, test, err):
        super().addFailure(test, err)
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        self.test_results.append({
            'name': test_name,
            'result': 'FAIL',
            'details': str(err[1]),
            'input_data': input_data,
            'expected_data': expected_data
        })

class CustomTestRunner:
    def __init__(self, stream=None, verbosity=1):
        self.stream = stream if stream else sys.stdout
        self.verbosity = verbosity
        
    def run(self, test):
        result = CustomTestResult(self.stream, self.verbosity)
        test(result)
        self.print_results(result)
        self.export_to_excel(result)
        return result
    
    def print_results(self, result):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("\n===== KẾT QUẢ TEST =====")
        print(f"Thời gian chạy: {current_time}\n")
        for idx, test_result in enumerate(result.test_results, 1):
            print(f"--- Test {idx} ---")
            print(f"Test Case        : {test_result['name']}")
            print(f"Kết quả          : {test_result['result']}")
            print(f"Dữ liệu đầu vào  : {test_result['input_data']}")
            print(f"Kết quả mong đợi : {test_result['expected_data']}")
            if test_result['details']:
                print(f"Chi tiết         : {test_result['details']}")
            print("-" * 40)

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Thông tin chung
        ws["A1"] = "Function Code"
        ws["B1"] = "dangnhap_admin"
        ws["A2"] = "Function Name"
        ws["B2"] = "Đăng nhập admin"
        ws["A3"] = "Created By"
        ws["B3"] = "Tên bạn"
        ws["A4"] = "Test requirement"
        ws["B4"] = "Kiểm tra chức năng đăng nhập admin với các trường hợp đầu vào"
        ws["A6"] = "STT"
        ws["B6"] = "Test Case"
        ws["C6"] = "Dữ liệu đầu vào"
        ws["D6"] = "Kết quả mong đợi"
        ws["E6"] = "Kết quả thực tế"
        ws["F6"] = "Trạng thái"

        # Định dạng header
        for col in "ABCDEF":
            ws[f"{col}6"].font = Font(bold=True)
            ws[f"{col}6"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col}6"].fill = PatternFill(start_color="003399FF", end_color="003399FF", fill_type="solid")
            ws[f"{col}6"].font = Font(color="FFFFFF", bold=True)

        # Ghi từng test case
        for idx, r in enumerate(result.test_results, 1):
            ws.append([
                idx,
                r['name'],
                r['input_data'],
                r['expected_data'],
                r['details'] if r['details'] else r['expected_data'],
                r['result']
            ])

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save("test_results.xlsx")
        print("Đã xuất kết quả ra file test_results.xlsx")

app = QApplication(sys.argv)

class Testdangnhap_admin(unittest.TestCase):
    def setUp(self):
        self.dn = dangnhap_admin()

    def tearDown(self):
        pass

    @patch("index.dangnhap_admin.QMessageBox")
    def test_tt_rong(self, mock_msgbox):
        self._input_data = "Tài khoản: '', Mật khẩu: ''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        self.dn.txt_ten_tk.setText("")
        self.dn.txt_mat_khau.setText("")
        self.dn.login()
        mock_msgbox.warning.assert_called_once_with(
            self.dn, "Lỗi", "Vui lòng nhập đầy đủ thông tin!"
        )

    @patch("index.dangnhap_admin.QMessageBox")
    def test_tt_thieu_mk(self, mock_msgbox):
        self._input_data = "Tài khoản: 'qqq@gmail.com', Mật khẩu: ''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        self.dn.txt_ten_tk.setText("qqq@gmail.com")
        self.dn.txt_mat_khau.setText("")
        self.dn.login()
        mock_msgbox.warning.assert_called_once_with(
            self.dn, "Lỗi", "Vui lòng nhập đầy đủ thông tin!"
        )

    @patch("index.dangnhap_admin.QMessageBox")
    def test_tt_thieu_tk(self, mock_msgbox):
        self._input_data = "Tài khoản: '', Mật khẩu: '123'"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        self.dn.txt_ten_tk.setText("")
        self.dn.txt_mat_khau.setText("123")
        self.dn.login()
        mock_msgbox.warning.assert_called_once_with(
            self.dn, "Lỗi", "Vui lòng nhập đầy đủ thông tin!"
        )

    @patch("index.dangnhap_admin.QMessageBox")
    def test_sai_dinhdang(self, mock_msgbox):
        self._input_data = "Tài khoản: 'qqq@gmail', Mật khẩu: '1234'"
        self._expected_data = "Hiển thị lỗi sai định dạng email"
        self.dn.txt_ten_tk.setText("qqq@gmail")
        self.dn.txt_mat_khau.setText("1234")
        self.dn.login()
        mock_msgbox.warning.assert_called_once_with(
            self.dn, "Lỗi", "Tên tài khoản phải có đuôi @gmail.com"
        )

    @patch("index.dangnhap_admin.QMessageBox")
    def test_sai_mk(self, mock_msgbox):
        self._input_data = "Tài khoản: 'qqq@gmail.com', Mật khẩu: '1234123'"
        self._expected_data = "Hiển thị lỗi đăng nhập thất bại"
        self.dn.txt_ten_tk.setText("qqq@gmail.com")
        self.dn.txt_mat_khau.setText("1234123")
        self.dn.login()
        mock_msgbox.warning.assert_called_once_with(
            self.dn, "Login output", "Đăng nhập thất bại"
        )

if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(Testdangnhap_admin)
    runner = CustomTestRunner(verbosity=2)
    result = runner.run(suite)