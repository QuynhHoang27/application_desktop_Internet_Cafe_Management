import sys
import unittest
from unittest.mock import patch, MagicMock
from PyQt6.QtWidgets import QApplication
from index.taikhoan_nhanvien import taikhoan_nhanvien
from datetime import datetime, timedelta
import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class CustomTestResult(unittest.TestResult):
    def __init__(self, stream=None, verbosity=1):
        super().__init__()
        self.stream = stream
        self.verbosity = verbosity
        self.test_results = []
        
    def startTest(self, test):
        super().startTest(test)
        
    def addSuccess(self, test):
        super().addSuccess(test)
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        self.test_results.append({
            'name': test_name,
            'result': 'PASS',
            'details': '',
            'input_data': input_data,
            'expected_data': expected_data
        })
        
    def addError(self, test, err):
        super().addError(test, err)
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        details = f"Lỗi hệ thống hoặc ngoại lệ: {str(err[1])}"
        self.test_results.append({
            'name': test_name,
            'result': 'ERROR',
            'details': details,
            'input_data': input_data,
            'expected_data': expected_data
        })
        
    def addFailure(self, test, err):
        super().addFailure(test, err)
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, '_input_data', '')
        expected_data = getattr(test, '_expected_data', '')
        err_msg = str(err[1])
        
        # Chuyển lỗi sang tiếng Việt cho các trường hợp thường gặp
        if "assertIn" in err_msg and "Vui lòng nhập đầy đủ thông tin" in err_msg:
            details = "Vui lòng nhập đầy đủ thông tin!"
        elif "assertIn" in err_msg and "Tên tài khoản phải có đuôi @gmail.com" in err_msg:
            details = "Tên tài khoản phải có đuôi @gmail.com!"
        elif "assertIn" in err_msg and "đã tồn tại" in err_msg:
            details = "Tên tài khoản đã tồn tại!"
        elif "assertIn" in err_msg and "Thêm thông tin thành công" in err_msg:
            details = "Thêm thông tin thành công!!!!"
        elif "assertIn" in err_msg and "Cập nhập thông tin thành công" in err_msg:
            details = "Cập nhập thông tin thành công!"
        elif "assertIn" in err_msg and "Mã OTP không đúng" in err_msg:
            details = "Mã OTP không đúng!"
        else:
            details = f"Lỗi kiểm thử: {err_msg}"
        self.test_results.append({
            'name': test_name,
            'result': 'FAIL',
            'details': details,
            'input_data': input_data,
            'expected_data': expected_data
        })

class CustomTestRunner:
    def __init__(self, stream=None, verbosity=1):
        self.stream = stream if stream else sys.stdout
        self.verbosity = verbosity
        
    def run(self, test):
        result = CustomTestResult(self.stream, self.verbosity)
        test(result)
        self.print_results(result)
        self.export_to_excel(result)
        return result
    
    def print_results(self, result):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("\n===== KẾT QUẢ TEST =====")
        print(f"Thời gian chạy: {current_time}\n")
        for idx, test_result in enumerate(result.test_results, 1):
            print(f"--- Test {idx} ---")
            print(f"Test Case        : {test_result['name']}")
            print(f"Kết quả          : {test_result['result']}")
            print(f"Dữ liệu đầu vào  : {test_result['input_data']}")
            print(f"Kết quả mong đợi : {test_result['expected_data']}")
            if test_result['details']:
                print(f"Chi tiết         : {test_result['details']}")
            print("-" * 40)

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Thông tin chung
        ws["A1"] = "Function Code"
        ws["B1"] = "taikhoan_nhanvien"
        ws["A2"] = "Function Name"
        ws["B2"] = "Quản lý tài khoản nhân viên"
        ws["A3"] = "Created By"
        ws["B3"] = ""
        ws["A4"] = "Test requirement"
        ws["B4"] = "Kiểm tra validation các trường nhập thông tin tài khoản nhân viên"
        ws["A6"] = "STT"
        ws["B6"] = "Test Case"
        ws["C6"] = "Dữ liệu đầu vào"
        ws["D6"] = "Kết quả mong đợi"
        ws["E6"] = "Kết quả thực tế"
        ws["F6"] = "Trạng thái"

        # Định dạng header
        for col in "ABCDEF":
            ws[f"{col}6"].font = Font(bold=True, color="FFFFFF")
            ws[f"{col}6"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col}6"].fill = PatternFill(start_color="003399FF", end_color="003399FF", fill_type="solid")

        # Ghi từng test case
        for idx, r in enumerate(result.test_results, 1):
            ws.append([
                idx,
                r['name'],
                r['input_data'],
                r['expected_data'],
                r['details'] if r['details'] else r['expected_data'],
                r['result']
            ])

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save("taikhoan_nhanvien_test.xlsx")
        print("Đã xuất kết quả ra file taikhoan_nhanvien_test.xlsx")

app = QApplication(sys.argv)

class TestTaiKhoanNhanVien(unittest.TestCase):
    def setUp(self):
        # Mock database connection
        self.mock_conn = MagicMock()
        self.mock_cursor = MagicMock()
        
        with patch('mysql.connector.connect', return_value=self.mock_conn):
            self.form = taikhoan_nhanvien("admin_test")
            self.form.cursor = self.mock_cursor

    def tearDown(self):
        pass

    # ===== TEST THÊM TÀI KHOẢN NHÂN VIÊN =====
    
    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_them_thieu_thong_tin_ten(self, mock_msgbox):
        """Test thiếu tên tài khoản"""
        self._input_data = "Tên: '', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.form.txt_ten.setText("")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.them()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("Vui lòng nhập đầy đủ thông tin", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_them_thieu_thong_tin_matkhau(self, mock_msgbox):
        """Test thiếu mật khẩu"""
        self._input_data = "Tên: 'nhanvien@gmail.com', Mật khẩu: ''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.form.txt_ten.setText("nhanvien@gmail.com")
        self.form.txt_matkhau.setText("")
        
        self.form.them()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("Vui lòng nhập đầy đủ thông tin", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_ten_tai_khoan_khong_co_gmail(self, mock_msgbox):
        """Test tên tài khoản không có đuôi @gmail.com"""
        self._input_data = "Tên: 'nhanvien@yahoo.com', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi tên tài khoản phải có đuôi @gmail.com"
        
        self.form.txt_ten.setText("nhanvien@yahoo.com")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.them()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("Tên tài khoản phải có đuôi @gmail.com", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_ten_tai_khoan_khong_co_duoi_gmail(self, mock_msgbox):
        """Test tên tài khoản không có email"""
        self._input_data = "Tên: 'nhanvien123', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi tên tài khoản phải có đuôi @gmail.com"
        
        self.form.txt_ten.setText("nhanvien123")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.them()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("Tên tài khoản phải có đuôi @gmail.com", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_ten_tai_khoan_da_ton_tai(self, mock_msgbox):
        """Test tên tài khoản đã tồn tại"""
        self._input_data = "Tên: 'nhanvien@gmail.com', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi tên tài khoản đã tồn tại"
        
        # Mock cursor trả về kết quả đã tồn tại
        self.mock_cursor.fetchone.return_value = ("TKNV001", "nhanvien@gmail.com", "pass")
        
        self.form.txt_ten.setText("nhanvien@gmail.com")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.them()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("đã tồn tại", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QInputDialog")
    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_them_thanh_cong_voi_otp(self, mock_msgbox, mock_input):
        """Test thêm tài khoản thành công với OTP hợp lệ"""
        self._input_data = "Tên: 'newstaff@gmail.com', Mật khẩu: 'pass123', OTP: '123456'"
        self._expected_data = "Thêm thông tin thành công!!!!"
        
        # Mock các giá trị trả về
        self.mock_cursor.fetchone.side_effect = [None, ("TKNV001",)]  # Tên chưa tồn tại, lấy max ID
        self.form.tao_tu_dong_ma = MagicMock(return_value="TKNV002")
        self.form.send_otp = MagicMock(return_value=True)
        self.form.otp_code = "123456"
        self.form.otp_expiry = datetime.now() + timedelta(minutes=2)
        
        # Mock input dialog trả về OTP đúng
        mock_input.getText.return_value = ("123456", True)
        
        self.form.txt_ten.setText("newstaff@gmail.com")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.them()
        
        mock_msgbox.information.assert_called_once()
        self.assertIn("Thêm thông tin thành công", mock_msgbox.information.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QInputDialog")
    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_them_fail_otp_sai(self, mock_msgbox, mock_input):
        """Test thêm tài khoản fail vì OTP sai"""
        self._input_data = "Tên: 'newstaff2@gmail.com', Mật khẩu: 'pass123', OTP sai: '111111'"
        self._expected_data = "Mã OTP không đúng!"
        
        # Mock các giá trị trả về
        self.mock_cursor.fetchone.return_value = None  # Tên chưa tồn tại
        self.form.send_otp = MagicMock(return_value=True)
        self.form.otp_code = "123456"
        self.form.otp_expiry = datetime.now() + timedelta(minutes=2)
        
        # Mock input dialog trả về OTP SAI
        mock_input.getText.return_value = ("111111", True)
        
        self.form.txt_ten.setText("newstaff2@gmail.com")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.them()
        
        # Kiểm tra có gọi critical với message OTP sai
        mock_msgbox.critical.assert_called_once()
        self.assertIn("Mã OTP không đúng", mock_msgbox.critical.call_args[0][2])

    # ===== TEST CẬP NHẬT TÀI KHOẢN =====
    
    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_update_thieu_thong_tin(self, mock_msgbox):
        """Test cập nhật thiếu thông tin"""
        self._input_data = "Mã TK: 'TKNV001', Tên: '', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.form.txt_matk.setText("TKNV001")
        self.form.txt_ten.setText("")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.update()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("Vui lòng điền đầy đủ thông tin", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_update_ten_khong_co_gmail(self, mock_msgbox):
        """Test cập nhật tên không có @gmail.com"""
        self._input_data = "Mã TK: 'TKNV001', Tên: 'staff@yahoo.com', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi tên tài khoản phải có đuôi @gmail.com"
        
        self.form.txt_matk.setText("TKNV001")
        self.form.txt_ten.setText("staff@yahoo.com")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.update()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("Tên tài khoản phải có đuôi @gmail.com", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_update_ten_da_ton_tai(self, mock_msgbox):
        """Test cập nhật tên đã tồn tại"""
        self._input_data = "Mã TK: 'TKNV001', Tên: 'existing@gmail.com', Mật khẩu: 'pass123'"
        self._expected_data = "Hiển thị lỗi tài khoản đã tồn tại"
        
        # Mock cursor trả về tài khoản khác đã có tên này
        self.mock_cursor.fetchone.return_value = ("TKNV002",)
        
        self.form.txt_matk.setText("TKNV001")
        self.form.txt_ten.setText("existing@gmail.com")
        self.form.txt_matkhau.setText("pass123")
        
        self.form.update()
        
        mock_msgbox.warning.assert_called_once()
        self.assertIn("đã tồn tại", mock_msgbox.warning.call_args[0][2])

    @patch("index.taikhoan_nhanvien.QMessageBox")
    def test_update_thanh_cong(self, mock_msgbox):
        """Test cập nhật thành công"""
        self._input_data = "Mã TK: 'TKNV001', Tên: 'updated@gmail.com', Mật khẩu: 'newpass'"
        self._expected_data = "Cập nhập thông tin thành công"
        
        # Mock cursor không tìm thấy tên trùng
        self.mock_cursor.fetchone.return_value = None
        
        self.form.txt_matk.setText("TKNV001")
        self.form.txt_ten.setText("updated@gmail.com")
        self.form.txt_matkhau.setText("newpass")
        
        self.form.update()
        
        mock_msgbox.information.assert_called_once()
        self.assertIn("Cập nhập thông tin thành công", mock_msgbox.information.call_args[0][2])

if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestTaiKhoanNhanVien)
    runner = CustomTestRunner(verbosity=2)
    result = runner.run(suite)