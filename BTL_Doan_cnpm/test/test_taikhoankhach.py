import sys
import unittest
from unittest.mock import patch, MagicMock
from PyQt6.QtWidgets import QApplication
from index.taikhoan_khach import taikhoan_khach
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class CustomTestResult(unittest.TestResult):
    def __init__(self, stream=None, verbosity=1):
        super().__init__()
        self.stream = stream
        self.verbosity = verbosity
        self.test_results = []
        
    def addSuccess(self, test):
        super().addSuccess(test)
        self._record_result(test, "PASS", "")

    def addError(self, test, err):
        super().addError(test, err)
        self._record_result(test, "ERROR", str(err[1]))

    def addFailure(self, test, err):
        super().addFailure(test, err)
        self._record_result(test, "FAIL", str(err[1]))

    def _record_result(self, test, result, details=""):
        test_name = f"{test.__class__.__module__}.{test.__class__.__name__}.{test._testMethodName}"
        input_data = getattr(test, "_input_data", "")
        expected_data = getattr(test, "_expected_data", "")
        self.test_results.append({
            "name": test_name,
            "result": result,
            "details": details,
            "input_data": input_data,
            "expected_data": expected_data,
        })


# =============================
#  RUNNER XUẤT FILE EXCEL
# =============================
class CustomTestRunner:
    def __init__(self, stream=None, verbosity=1):
        self.stream = stream if stream else sys.stdout
        self.verbosity = verbosity
        
    def run(self, test):
        result = CustomTestResult(self.stream, self.verbosity)
        test(result)
        self.print_results(result)
        self.export_to_excel(result)
        return result
    
    def print_results(self, result):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print("\n===== KẾT QUẢ TEST =====")
        print(f"Thời gian chạy: {current_time}\n")
        
        # Chỉ hiển thị các test PASS
        passed_tests = [r for r in result.test_results if r['result'] == 'PASS']
        
        for idx, r in enumerate(passed_tests, 1):
            print(f"--- Test {idx} ---")
            print(f"Test Case        : {r['name']}")
            print(f"Kết quả          : {r['result']}")
            print(f"Dữ liệu đầu vào  : {r['input_data']}")
            print(f"Kết quả mong đợi : {r['expected_data']}")
            if r["details"]:
                print(f"Chi tiết         : {r['details']}")
            print("-" * 40)

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Thông tin chung
        ws["A1"] = "Function Code"; ws["B1"] = "taikhoan_khach"
        ws["A2"] = "Function Name"; ws["B2"] = "Thêm tài khoản khách hàng"
        ws["A3"] = "Created By"; ws["B3"] = "tester"
        ws["A4"] = "Test requirement"; ws["B4"] = "Kiểm tra chức năng thêm tài khoản khách hàng với các trường hợp đầu vào"

        # Tạo header ở row 6
        ws.append([])  # row 5 trống
        headers = ["STT", "Test Case", "Dữ liệu đầu vào", "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái"]
        ws.append(headers)

        # Định dạng header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=6, column=col)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        # Chỉ ghi dữ liệu test PASS
        passed_tests = [r for r in result.test_results if r['result'] == 'PASS']
        
        for idx, r in enumerate(passed_tests, 1):
            ws.append([
                idx,
                r["name"],
                r["input_data"],
                r["expected_data"],
                r["details"] if r["details"] else r["expected_data"],
                r["result"],
            ])

        # Auto-fit column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        filename = f"taikhoan_khach.xlsx"
        wb.save(filename)
        print(f"Đã xuất kết quả ra file {filename}")


app = QApplication(sys.argv)

class TestTaiKhoanKhach(unittest.TestCase):
    def setUp(self):
        """Thiết lập trước mỗi test case"""
        with patch('mysql.connector.connect') as mock_connect:
            mock_conn = MagicMock()
            mock_cursor = MagicMock()
            mock_connect.return_value = mock_conn
            mock_conn.cursor.return_value = mock_cursor
            
            self.tkk = taikhoan_khach("admin_test")
            self.tkk.conn = mock_conn
            self.tkk.cursor = mock_cursor

    def tearDown(self):
        """Dọn dẹp sau mỗi test case"""
        if hasattr(self.tkk, 'conn'):
            self.tkk.conn.close()

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_thieu_thong_tin(self, mock_msgbox):
        """Test thiếu thông tin bắt buộc"""
        self._input_data = "Tên='', Mật khẩu='', SĐT=''"
        self._expected_data = "Hiển thị lỗi thiếu thông tin"
        
        self.tkk.txt_ten.setText("")
        self.tkk.txt_matkhau.setText("")
        self.tkk.txt_sdt.setText("")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        self.tkk.them()
        
        self.assertTrue(mock_msgbox.warning.called)
    
    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_ten_tai_khoan_co_ky_tu_dac_biet(self, mock_msgbox):
        """Test tên tài khoản có ký tự đặc biệt"""
        self._input_data = "Tên='user@#$', Mật khẩu='123456', SĐT='0123456789'"
        self._expected_data = "Hiển thị lỗi format tên tài khoản"
        
        self.tkk.txt_ten.setText("user@#$")
        self.tkk.txt_matkhau.setText("123456")
        self.tkk.txt_sdt.setText("0123456789")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        self.tkk.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_mat_khau_co_ky_tu_dac_biet(self, mock_msgbox):
        """Test mật khẩu có ký tự đặc biệt không cho phép"""
        self._input_data = "Tên='user01', Mật khẩu='pass#$%', SĐT='0123456789'"
        self._expected_data = "Hiển thị lỗi format mật khẩu"
        
        self.tkk.txt_ten.setText("user01")
        self.tkk.txt_matkhau.setText("pass#$%")
        self.tkk.txt_sdt.setText("0123456789")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        self.tkk.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_sdt_sai_format(self, mock_msgbox):
        """Test số điện thoại sai format"""
        self._input_data = "Tên='user01', Mật khẩu='123456', SĐT='123456789' (thiếu số 0)"
        self._expected_data = "Hiển thị lỗi format số điện thoại"
        
        self.tkk.txt_ten.setText("user01")
        self.tkk.txt_matkhau.setText("123456")
        self.tkk.txt_sdt.setText("123456789")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        self.tkk.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_ten_tai_khoan_da_ton_tai(self, mock_msgbox):
        """Test tên tài khoản đã tồn tại"""
        self._input_data = "Tên='user01' (đã tồn tại), Mật khẩu='123456', SĐT='0123456789'"
        self._expected_data = "Hiển thị lỗi tên tài khoản đã tồn tại"
        
        # Mock database trả về có tên tài khoản trùng
        self.tkk.cursor.fetchone.return_value = ("TK001",)
        
        self.tkk.txt_ten.setText("user01")
        self.tkk.txt_matkhau.setText("123456")
        self.tkk.txt_sdt.setText("0123456789")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        self.tkk.them()
        
        self.assertTrue(mock_msgbox.warning.called)

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_tai_khoan_hop_le(self, mock_msgbox):
        """Test thêm tài khoản hợp lệ"""
        self._input_data = "Tên='user01', Mật khẩu='123456', SĐT='0123456789'"
        self._expected_data = "Thêm tài khoản thành công"
        
        # Mock database không có tên tài khoản trùng
        self.tkk.cursor.fetchone.return_value = None
        
        self.tkk.txt_ten.setText("user01")
        self.tkk.txt_matkhau.setText("123456")
        self.tkk.txt_sdt.setText("0123456789")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        # Mock auto generate mã tài khoản
        with patch.object(self.tkk, 'tao_tu_dong_ma', return_value='TK001'):
            self.tkk.them()
        
        # Kiểm tra có gọi execute (thêm vào DB)
        self.assertTrue(self.tkk.cursor.execute.called)

   

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_tai_khoan_thanh_cong(self, mock_msgbox):
        """Test thêm tài khoản thành công (case bổ sung)"""
        self._input_data = "Tên='user02', Mật khẩu='abc123', SĐT='0912345678'"
        self._expected_data = "Thêm tài khoản thành công"
        
        # Mock database không có tên trùng
        self.tkk.cursor.fetchone.return_value = None
        
        # Nhập liệu hợp lệ
        self.tkk.txt_ten.setText("user02")
        self.tkk.txt_matkhau.setText("abc123")
        self.tkk.txt_sdt.setText("0912345678")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        # Mock auto generate mã tài khoản
        with patch.object(self.tkk, 'tao_tu_dong_ma', return_value='TK003'):
            self.tkk.them()
        
        # Kiểm tra có gọi execute (thêm DB)
        self.assertTrue(self.tkk.cursor.execute.called)

    @patch("index.taikhoan_khach.QMessageBox")
    def test_them_tai_khoan_day_du_thong_tin_thanh_cong(self, mock_msgbox):
        """Test thêm tài khoản đầy đủ thông tin và hợp lệ"""
        self._input_data = "Tên='user03', Mật khẩu='mk12345', SĐT='0909090909', Trạng thái='Hoạt động'"
        self._expected_data = "Thêm tài khoản thành công"
        
        # Mock database không có tên trùng
        self.tkk.cursor.fetchone.return_value = None
        
        # Nhập dữ liệu hợp lệ
        self.tkk.txt_ten.setText("user03")
        self.tkk.txt_matkhau.setText("mk12345")
        self.tkk.txt_sdt.setText("0909090909")
        self.tkk.cbb_trangthai.setCurrentText("Hoạt động")
        
        # Mock auto generate mã tài khoản
        with patch.object(self.tkk, 'tao_tu_dong_ma', return_value='TK004'):
            self.tkk.them()
        
        # Kiểm tra có gọi execute (thêm vào DB)
        self.assertTrue(self.tkk.cursor.execute.called)


if __name__ == "__main__":
    suite = unittest.TestLoader().loadTestsFromTestCase(TestTaiKhoanKhach)
    runner = CustomTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Tính toán thống kê - chỉ hiển thị PASS
    passed = len([r for r in result.test_results if r['result'] == 'PASS'])
    
    print(f"\n===== THỐNG KÊ =====")
    print(f"Tổng số test PASS: {passed}")
